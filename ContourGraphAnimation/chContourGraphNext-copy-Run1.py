#! /usr/bin/env python2.7
# -*- coding: utf-8 -*-
"""
Created on Sun Mar 08 20:23:19 2015
Змінено: 24.04.2018 на основі chContourGraphNext-copy із GoogleDive/storage (2)

@author: Stepan
"""
from __future__ import division

import sys, os
import gc
import subprocess
import shutil
from datetime import timedelta
from collections import OrderedDict

import re
re_pattern_sheet_data = re.compile(r'([A-Za-z]+)([0-9]+[.,]?[0-9]*)')

import matplotlib as mpl
import numpy as np
import scipy.ndimage    # Рівномірна інтерполяція зображень.
import matplotlib.cm as cm  # Кольорові мапи.
import matplotlib.mlab as mlab
import matplotlib.pyplot as plt

# Робота з файлами електронних таблиць формату XLSX.
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import Worksheet, cells_from_range

sys.path.append(r'C:\Users\ichet\Dropbox\MyApps\chFramevork')
from chCore.chRun import ChApp
from chCore.ch_setupcon import setup_console

mpl.rcParams['font.family'] = 'serif'   # Сімейство шрифтів з зарубками.
mpl.rcParams['font.serif'] = 'Times New Roman', 'Ubuntu', 'Arial', \
                               'Tahoma','Calibri'
mpl.rcParams['axes.titlesize'] = 'medium'
# Виокристовувати регіональні налаштування виводу даних.
mpl.rcParams['axes.formatter.use_locale'] = True

# Налаштування області виводу графіку.
mpl.rcParams['figure.subplot.bottom'] = 0.115
mpl.rcParams['figure.subplot.left'] = 0.12
mpl.rcParams['figure.subplot.right'] = 0.98
mpl.rcParams['figure.subplot.top'] = 0.93

mpl.rcParams['legend.fontsize'] = 'medium'

mpl.rcParams['xtick.direction'] = 'out'
mpl.rcParams['ytick.direction'] = 'out'

setup_console()

fig_data = dict(figsize=(8, 6), dpi=100)

# ==============================================================================
def create_animation(xlsx_file, sheet_name, cells_range, total_time, main_title, lower_title,
                     smoothing=10, fps=24, rectime_secs=60, is_grid=False, is_contour=False):
    ur'''Створює анімацію
    version: 0.1.0
    args:
    - xlsx_file(basestring) - назва XLSX-файлу;
    - sheet_name(basestring, int) - назва або індекс аркуша в робочій книзі;
    - cells_range(tuple) - діапазон клітинок, де розміщуються дані (тіло таблиці 
      без боковиків);
    - title(basestring) - назва графіку;
    - smoothing(int) - кількість рівнів згладжування;
    - fps(int) - частота кадрів (кількість кадрів за секунду);
    - rectime_secs(int) - тривалість створеного відео.
    '''
    
    #try:
     #   shutil.rmtree(dirout_with_file, ignore_errors=False, onerror=None)
    #except WindowsError:
     #   print u'''Якийсь процес використовує файли в папці "%s". 
#Закрийте всі программи, в яких відкриті або використовуються файли в папці, і \
#повторите попытку!''' % os.path.basename(dirout_with_file).encode('utf-8'))
 #       continue
    
    # Завантаження даних з XLSX-аркуша.
    wb = load_workbook(xlsx_file)
    
    # Визначення робочого аркуша по індексу або імені.
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[sheet_name] if isinstance(
      sheet_name, int) else sheet_name)
    
    #print ws.title
    #print os.path.basename(xlsx_file).encode('cp866')
    print u'Відкриваю аркуш "%s" робочої книги "%s"' % (ws.title, 
                                     os.path.basename(xlsx_file))
    
    # Масив осей.
    x = np.array([0, 10, 20, 30, 40, 50])
    y = np.array([0, 10, 20, 30, 40])
    
    def extvec(vector, di=1):
        new_vector = []
        previous_value = vector[0]
        
        for value in vector[1:]:
            new_vector += np.linspace(previous_value, value, di, False).tolist()
            previous_value = value
        
        new_vector.append(value)
        
        return np.array(new_vector)
    
    X = extvec(x)
    Y = extvec(y)
    
    # Розширення масиву в smoothing раз в кожному напрямку завдяки інтерполяції 
    # проміжних значень.
    if smoothing:
        X = scipy.ndimage.zoom(X, smoothing)
        Y = scipy.ndimage.zoom(Y, smoothing)
    
    #X = []
    #Y = []
    T = []
    # Загальна кількість кадрів за всю тривалість відео.
    total_frames = fps * rectime_secs
    
    for column_range in cells_range:
        t = []
        
        for c in cells_from_range(column_range):
            t.append(ws.cell(c[0]).value)
        
        # Переформатування масиву.
        t = np.array(t).reshape(5, 6)
        # Розширення матриці в smoothing раз в кожному напрямку завдяки 
        # інтерполяції проміжних значень.
        if smoothing:
            t = scipy.ndimage.zoom(t, smoothing)
        T.append(t)
    
    T = np.array(T)
    #print T.shape
    N = total_frames // len(T)
    #print N
    #print T.shape
    tot_T =  scipy.ndimage.zoom(T, [N,1,1])
    #print scipy.ndimage.zoom(T, 500).shape
    #print scipy.ndimage.zoom(T[0:2], total_frames/(1.*len(T[0]))).shape
    #print total_frames
    #T = scipy.ndimage.zoom(T, total_frames)
    #print T.shape
    
    Y = Y[: : -1]
    
    frame_fname_format = '%0' + str(len(str(len(tot_T)))) + 'd.png'
    print u'Всього створюється %d кадрiв формою %dx%d точок...' \
      % (len(tot_T), T[0].shape[0], T[0].shape[1])
    
    progress_char = '#'
    progress_width = 50
    
    fig = plt.figure(**fig_data)
    #fig.hold(False)
    
    font = lambda fs: mpl.font_manager.FontProperties(family='Times New Roman', 
      style='normal', size=fs, weight='normal', stretch='normal')
    
    T_min = tot_T.min()
    T_max = tot_T.max()
    cb_k = 5
    cb_value_min = divmod(T_min, cb_k)[0] * cb_k
    cb_value_max = divmod(T_max, cb_k)[0] * (cb_k + 1)
    cb_ticks = np.arange(cb_value_min, cb_value_max+cb_k, cb_k)
    norm = mpl.colors.Normalize(vmin=cb_value_min, vmax=cb_value_max)
    
    print u'Запуск процесу...'
    
    for i, T_i in enumerate(tot_T):
        fig.clear()
        try:
            del ax
        except UnboundLocalError:
            pass
        ax = fig.add_subplot(111)
        ax.figure.subplots_adjust(left=.01, bottom = .16, top=.92, right=.92)
        
        frame_i = i + 1
        #CS = plt.contour(x, y, t)
        
        #CS = plt.contour(X, Y, T, 15, nchunk=5, linewidths=1.5)
        #plt.clabel(CS, inline=True, fmt='%.1f')
        
        if is_contour:
            CS = plt.contour(X, Y, T_i, linewidths=1.5, colors='k') #, nchunk=5
            fmt = dict()
            for lbl in CS.levels:
                fmt[lbl] = (u'{:.2f}'.format(lbl)).replace('.', ',')
            ax.clabel(CS, CS.levels, inline=True, fmt=fmt)
            for cont_label in CS.labelTexts:
                # Всі надписи - чорні.
                #if discolor: cont_label.set_color('k')
                
                cont_label.set_fontproperties(font(14))
        
        im = ax.imshow(T_i, interpolation='bilinear', origin='upper', extent=(0, 50, 0, 40), cmap='jet')
        
        ax1 = fig.add_axes([0.9, 0.16, 0.025, 0.92-.16])
        cb = mpl.colorbar.ColorbarBase(ax1, cmap='jet',
                                norm=norm,
                                orientation='vertical')
        cb.set_label(ur'Температура на поверхні, ${}^{\mathrm{\circ}}$C', fontproperties=font(14))
        #cb = fig.colorbar(im, anchor=(.7, .35), ticks=cb_ticks)
        #cb.set_clim(cb_T_min, cb_T_max)
        #cb.ax.set_yticks(cb_ticks)
        #cb.ax.set_yticklabels([str(x) for x in cb_ticks])#(cb_T_min, cb_T_max)
        #plt.title(title, fontsize=14)
        # Назва графіку.
        title = ax.set_title(main_title, ha='center', 
                             fontproperties=font(14))
        title.set_position([title.get_position()[0], -0.125])
        
        # Нижня назва графіку.
        lwr_title = ax.text(title.get_position()[0], -.15, lower_title, 
          ha='center', va='top', fontproperties=font(14), transform=ax.transAxes)
        
        
        ax.set_xticklabels(['0', '0,1', '0,2', '0,3', '0,4', '0,5'], 
                           fontproperties=font(14))
        ax.set_yticks([0, 10, 20, 30, 40])
        ax.set_yticklabels(['0', '0,1', '0,2', '0,3', '0,4'],
                           fontproperties=font(14))
        # Мітка осі x.
        xl = ax.set_xlabel(u'a, м', ha='left', 
          va='top', fontproperties=font(14))
        ticklab = ax.xaxis.get_ticklabels()[0]
        trans = ticklab.get_transform()
        ax.xaxis.set_label_coords(52, 0, transform=trans)
        
        yl = ax.set_ylabel(u'b, м', rotation=0, 
          ha='right', va='bottom',
          fontproperties=font(14))
        ticklab = ax.yaxis.get_ticklabels()[1]
        trans = ticklab.get_transform()
        ax.yaxis.set_label_coords(0, 42, 
          transform=trans)
            
        al = 7 # arrow length in points
        arrowprops=dict(#clip_on=False, # plotting outside axes on purpose
          frac=1., # make end arrowhead the whole size of arrow
          headwidth=al, # in points
          facecolor='g')
        kwargs = dict(  
                      xycoords='axes fraction',
                      textcoords='axes fraction',
                      arrowprops= dict(arrowstyle="->"),
                      size=20.
                   )
        
        # Розміщення міток підписів осей.
        ax.annotate("",xy=(1.085,0),xytext=(1, 0), **kwargs) # bottom spine arrow
        ax.annotate("",(0,1.085),xytext=(0, 1), **kwargs) # left spin arrow
        
        # Видимість сітки.
        if is_grid:
            ax.grid(b=True, color='#00ff00',
                    linestyle='-.',
                    linewidth=.5)
        else: ax.grid(False)
        
        fig.text(.5, .93, ur'Пройдений час: $\mathrm{{\tau}}$ = {}'.format(timedelta(
                   seconds=round((i+1)/fps * total_time*60/rectime_secs))),
                   fontweight='bold', fontproperties=font(14), ha='center',
                   va='bottom', transform=ax.transAxes)
        
        plt.savefig(os.path.join(workdir, frame_fname_format % frame_i), 
                    dpi=fig_data['dpi'], format='png')
        # Очищає пам'ять.
        ##ax.figure.clf()
        ##plt.close()
        ##gc.collect()
        
        fraction = round(frame_i / len(tot_T), 2)
        progress_line = progress_char * int(round(fraction * progress_width, 0))
        
        progress_line = u'Створено %d кадрiв: %s' % (frame_i, 
          '%s %d%%' % (progress_line, 100*fraction))
        sys.stdout.write(progress_line)
        sys.stdout.flush()
        sys.stdout.write('\b' * len(progress_line))
    
    sys.stdout.flush()
    video_settings = {'vcodec': 'mpeg4', 'vtag': 'xvid', 'vqscale': 3,
                      'acodec': 'libmp3lame', 'aqscale': 4, 'container': 'avi'}
    print u'Створення відео з налаштуваннями {}: ...'.format(', '.join(
      '{} = {}'.format(key, value) for key, value in video_settings.items()))
    path_ffmpeg = ur'ffmpeg-20180424-d9706f7-win64-static\bin\ffmpeg.exe'
    files_imresc = os.path.join(workdir, frame_fname_format)
    video_file = os.path.join(workdir, 'animation.'+video_settings['container'])
    
    try:
        os.remove(video_file)
    except WindowsError:
        pass
    
    command = (ur'{ffmpeg} -framerate {framerate} -i {files_imresc} '
                     ur'-c:v {vcodec} -vtag {vtag} -qscale:v {vqscale} '
                     ur'-c:a {acodec} -qscale:a {aqscale} {video_file}'.format(
                     ffmpeg=path_ffmpeg, framerate=fps, files_imresc=files_imresc,
                     video_file=video_file, **video_settings))
    process = subprocess.Popen(command, stdout=subprocess.PIPE,
                               stderr=subprocess.STDOUT, bufsize=1, universal_newlines=True)
    while True:
        line = process.stdout.readline()
        if line != b'':
            os.write(1, line)
        else:
            break
    
    sys.stdout.flush()
    print u'Процес завершено!'

#app = ChApp(product_id='c09a7d54-11e4-58b6-8aed-ab409a36a76a', 
            #product_name='ContourGraph')
#app.start()
#ita = app.calc(idata = OrderedDict([
#  (u'Повна назва робочої книги електронних таблиць, включаючи розширення і шлях до XLSX-файлу:', 
#    (basestring, u'Досліди на 21.03.2015 21-31\\doslidu-heliostina.M1\\CIRKULACIJA\\Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx')), 
#  (u'Назва робочого аркуша:', (basestring, u'd10l10x20v0,15V15лI600')), 
#  (u'Масив стовпців з діапазоном клітинок даних:' (tuple, ('C4:C33', 'D4:D33', 
#    'E4:E33', 'F4:F33', 'G4:G33', 'H4:H33', 'I4:I33'))), 
#  (u'Назва графіку:', (basestring, ur'Циркуляція $d = 10$ мм, $s = 10$ мм, $\delta = 20$ мм, $v = 0,15$ м/с')), 
#  ('Множник розширення:', (int, 10)), 
#  (u'Частота кадрів (кількість кадрів за секунду):', (int, 24)), 
#  (u'Тривалість відео в секундах:', (int, 60)),
                            # ]), )
#ita.efunc = create_animation

import re
re_pattern_sheet_data = re.compile(r'([A-Za-z]+)([0-9]+[.,]?[0-9]*)')
def get_title(sheet_name):
    sheet_name_data = dict(re_pattern_sheet_data.findall(sheet_name))
    if 'l' in sheet_name_data.keys():
        sheet_name_data['L'] = sheet_name_data.pop('l')
    return (ur'd = {d} мм, L = {L} мм, x = {x} см, v = {v} л/(хв$\cdot$м${{}}^2$), '
                ur'V = {V} м${{}}^3$, I = {I} Вт/м${{}}^2$'.format(
                    **sheet_name_data) )

fps = 24
video_time = 20 # сек
smoothing = 10
is_contour = False

i = 1
if i == 0:
    DIR_WORK = r"C:\Users\ichet\Dropbox\myDocs\St"
    workdir = os.path.join(DIR_WORK, r'ContourGraphAnimation\Anim_GREB')
    fname = os.path.join(DIR_WORK, u'Гребінка.М1Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx')
    sheet_name = u'd10L20x10v0,3V15I600'
    cells_range = ('B2:B31', 'C2:C31', 'D2:D31', 
      'E2:E31')
    
    total_time = 90 # хв
    title = u'Розподілення температур в геліостіні з гребінкою'
    subtitle = u'{}'.format(get_title(sheet_name))
else:
    DIR_WORK = r"C:\Users\ichet\Dropbox\myDocs\St"
    workdir = os.path.join(DIR_WORK, r'ContourGraphAnimation\Anim_ZMIJ')
    fname = os.path.join(DIR_WORK, u'Змієвик.М1Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx')
    sheet_name = u'd10l10x20v0,25V15лI600'
    cells_range=('B3:B32', 'C3:C32', 'D3:D32', 
      'E3:E32', 'F3:F32', 'G3:G32')
    
    total_time = 100
    title = u'Розподілення температур у геліостіні зі змійовиком'
    subtitle = u'{}'.format(get_title(sheet_name))

create_animation(fname, sheet_name, cells_range, total_time, 
                 main_title = title, lower_title=subtitle, smoothing=smoothing, fps=fps,
                 rectime_secs=video_time, is_contour=is_contour)

raise SystemExit(0)
for i, (range, title) in enumerate((('D4:D33', title(20)), ('E4:E33', title(40)), 
  ('F4:F33', title(60)), ('G4:G33', title(80)), ('H4:H33', title(100)), 
  ('I4:I33', title(120)))):
    
    gen_contour(fname, sheet_name, range, title, smoothing)
    name = ('fig_img%d.png' % i) if smoothing > 1 else ('fig 0%d.png' % i+1)
    plt.savefig(os.path.join(workdir, name), format='png')

plt.show()
