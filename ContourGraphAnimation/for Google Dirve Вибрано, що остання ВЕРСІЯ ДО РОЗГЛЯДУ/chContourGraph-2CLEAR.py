#! /usr/bin/env python2.7
# -*- coding: utf-8 -*-
"""
Created on Sun Mar 08 20:23:19 2015

@author: Stepan
"""
from __future__ import division

import sys, os
import shutil
from collections import OrderedDict

import matplotlib as mpl
import numpy as np
import scipy.ndimage    # Рівномірна інтерполяція зображень.
import matplotlib.cm as cm  # Кольорові мапи.
import matplotlib.mlab as mlab
import matplotlib.pyplot as plt

# Робота з файлами електронних таблиць формату XLSX.
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import Worksheet, cells_from_range

#from chCore.chRun import ChApp

mpl.rcParams['font.family'] = 'serif'   # Сімейство шрифтів з зарубками.
mpl.rcParams['font.serif'] = 'Times New Roman', 'Ubuntu', 'Arial', \
                               'Tahoma','Calibri'
mpl.rcParams['axes.titlesize'] = 'medium'
# Виокристовувати регіональні налаштування виводу даних.
mpl.rcParams['axes.formatter.use_locale'] = True

# Налаштування області виводу графіку.
mpl.rcParams['figure.subplot.bottom'] = 0.115
mpl.rcParams['figure.subplot.left'] = 0.12
mpl.rcParams['figure.subplot.right'] = 0.98
mpl.rcParams['figure.subplot.top'] = 0.93

mpl.rcParams['legend.fontsize'] = 'medium'

mpl.rcParams['xtick.direction'] = 'out'
mpl.rcParams['ytick.direction'] = 'out'

# ==============================================================================
def create_contour(xlsx_file, sheet_name, cells_range, title, smoothing=10):
    ur'''Створює анімацію
    version: 0.1.0
    args:
    - xlsx_file(basestring) - назва XLSX-файлу;
    - sheet_name(basestring, int) - назва або індекс аркуша в робочій книзі;
    - cells_range(tuple) - діапазон клітинок, де розміщуються дані (тіло таблиці 
      без боковиків);
    - title(basestring) - назва графіку;
    - smoothing(int) - кількість рівнів згладжування.
    '''
    
    #try:
     #   shutil.rmtree(dirout_with_file, ignore_errors=False, onerror=None)
    #except WindowsError:
     #   print u'''Якийсь процес використовує файли в папці "%s". 
#Закрийте всі программи, в яких відкриті або використовуються файли в папці, і \
#повторите попытку!''' % os.path.basename(dirout_with_file).encode('utf-8'))
 #       continue
    
    # Завантаження даних з XLSX-аркуша.
    wb = load_workbook(xlsx_file)
    
    # Визначення робочого аркуша по індексу або імені.
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[sheet_name] if isinstance(
      sheet_name, int) else sheet_name)
    
    #print ws.title
    #print os.path.basename(xlsx_file).encode('cp866')
    #print u'Відкриваю аркуш "%s" робочої книги "%s"' % (ws.title, 
     #                                os.path.basename(xlsx_file))
    
    # Масив осей.
    x = np.array([0, 10, 20, 30, 40, 50])
    y = np.array([0, 10, 20, 30, 40])
    
    # Розширення масиву в smoothing раз в кожному напрямку завдяки інтерполяції 
    # проміжних значень.
    if smoothing:
        x = scipy.ndimage.zoom(x, smoothing, order=3)
        y = scipy.ndimage.zoom(y, smoothing, order=3)
    
    T = []
    
    for column_range in cells_range:
        t = []
        
        for c in cells_from_range(column_range):
            t.append(float(ws.cell(c[0]).value))
        
        # Переформатування масиву.
        t = np.array(t).reshape(5, 6)
        # Розширення матриці в smoothing раз в кожному напрямку завдяки 
        # інтерполяції проміжних значень.
        if smoothing: t = scipy.ndimage.zoom(t, smoothing, order=3)
       # T.append(t)
    
    #T = np.array(T)
    #print T.shape
    #print N
    #print T.shape
    #print scipy.ndimage.zoom(T, 500).shape
    #print scipy.ndimage.zoom(T[0:2], total_frames/(1.*len(T[0]))).shape
    #print total_frames
    #T = scipy.ndimage.zoom(T, total_frames)
    #print T.shape
    
    #y = y[: : -1]
    
    #fig = plt.figure()
    #fig.hold(False)
    CS = plt.contour(x, y, t, 15, nchunk=5, linewidths=1.5)
    #plt.clabel(CS, inline=True, fmt='%.1f')
    plt.show()
    
    #for i, T_i in enumerate(tot_T):
    #    fig.clear()
        #CS = plt.contour(x, y, t)
        
        #CS = plt.contour(X, Y, T, 15, nchunk=5, linewidths=1.5)
        #plt.clabel(CS, inline=True, fmt='%.1f')
        ###plt.imshow(T_i, interpolation='bilinear', origin='upper', extent=(0, 50, 40, 0))
        ###fig.savefig(os.path.join(workdir, 'fig_img%d.png' % frame_i), format='png')
        ###plt.title(title, fontsize=14)
        
xlsx_file = u'Досліди на 21.03.2015 21-31\\doslidu-heliostina.M1\\CIRKULACIJA\\Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx'
sheet_name = u'd10l10x20v0,15V15лI600'
cells_range=('D4:D33',)
title = 'z'
create_contour(xlsx_file, sheet_name, cells_range, title, smoothing=10)
raise SystemExit
app = ChApp(product_id='c09a7d54-11e4-58b6-8aed-ab409a36a76a', 
            product_name='ContourGraph')
app.start()
ita = app.calc(idata = OrderedDict([
  (u'Повна назва робочої книги електронних таблиць, включаючи розширення і шлях до XLSX-файлу:', 
    (basestring, u'Досліди на 21.03.2015 21-31\\doslidu-heliostina.M1\\CIRKULACIJA\\Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx')), 
  (u'Назва робочого аркуша:', (basestring, u'd10l10x20v0,15V15лI600')), 
  (u'Масив стовпців з діапазоном клітинок даних:' (tuple, ('C4:C33', 'D4:D33', 
    'E4:E33', 'F4:F33', 'G4:G33', 'H4:H33', 'I4:I33'))), 
  (u'Назва графіку:', (basestring, ur'Циркуляція $d = 10$ мм, $s = 10$ мм, $\delta = 20$ мм, $v = 0,15$ м/с')), 
  ('Множник розширення:', (int, 10)), 
  (u'Частота кадрів (кількість кадрів за секунду):', (int, 24)), 
  (u'Тривалість відео в секундах:', (int, 60)),
                             ]), )
ita.efunc = create_animation
#workdir = ur'Graphics1'
#fname = u'Досліди на 21.03.2015 21-31\\doslidu-heliostina.M1\\CIRKULACIJA\\Режим.Циркуляції.Розподіл.тем.на пов.V15I15.xlsx'
#sheet_name = ur'd10l10x20v0,15V15лI600'
#title = lambda text: ur'Циркуляція $d = 10$ мм, $s = 10$ мм, $\delta = 20$ мм, $v = 0,15$ м/с'
#smoothing = 10

#create_animation(fname, sheet_name, cells_range=('C4:C33', 'D4:D33', 'E4:E33', 
 # 'F4:F33', 'G4:G33', 'H4:H33', 'I4:I33'), title = title, smoothing=smoothing, 
 # fps=24, rectime_secs=60)

raise SystemExit
for i, (range, title) in enumerate((('D4:D33', title(20)), ('E4:E33', title(40)), 
  ('F4:F33', title(60)), ('G4:G33', title(80)), ('H4:H33', title(100)), 
  ('I4:I33', title(120)))):
    
    gen_contour(fname, sheet_name, range, title, smoothing)
    name = ('fig_img%d.png' % i) if smoothing > 1 else ('fig 0%d.png' % i+1)
    plt.savefig(os.path.join(workdir, name), format='png')

plt.show()
